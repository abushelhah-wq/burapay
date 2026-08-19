#!/usr/bin/env python3
"""
Build results/gateway_benchmark_workbook.xlsx.

Run with no arguments and it produces the documentation-based workbook: the
call-count comparison, the per-gateway endpoint map, and the 22 open data gaps,
plus timing tabs with headers ready to receive real numbers.

Point it at a BuraPay JSON export and it folds the measured timings into those tabs
and adds a documented-vs-measured sheet.

    python build_workbook.py                        # uses results/latest.json if present
    python build_workbook.py --results path.json    # a specific run
    python build_workbook.py --docs-only            # ignore any measured data
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Sequence

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(SCRIPT_DIR)
for path in (SCRIPT_DIR, ROOT):
    if path not in sys.path:
        sys.path.insert(0, path)
RESULTS_DIR = os.path.join(ROOT, "results")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guard
    print("openpyxl is not installed.\n"
          "  Fix: python3 -m pip install -r requirements.txt --break-system-packages",
          file=sys.stderr)
    raise SystemExit(1)

from reference_data import (CALL_COUNT_MATRIX, DATA_GAPS, ENDPOINT_MAP,  # noqa: E402
                            GATEWAYS, KEY_FINDINGS, METHODOLOGY, SOURCES)

#: The operations this documentation workbook compares. Declared here rather than
#: imported from the application: this script documents what the vendors *say*, and
#: it must keep working unchanged even as the platform's own flow list evolves.
FLOWS = ("hosted_checkout", "direct_api", "capture", "refund", "void", "mit")

FLOW_LABELS = {
    "hosted_checkout": "Hosted checkout (redirect)",
    "direct_api": "Direct API (server-to-server)",
    "capture": "Capture",
    "refund": "Refund",
    "void": "Void",
    "mit": "MIT / recurring charge",
}

# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #

TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SUBHEADER_FONT = Font(bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
BASELINE_FILL = PatternFill("solid", fgColor="FFF2CC")   # Geidea column
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")       # unverified / open gap
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title(sheet, text: str, width: int) -> int:
    sheet.cell(row=1, column=1, value=text).font = TITLE_FONT
    sheet.cell(row=1, column=1).fill = TITLE_FILL
    for column in range(2, width + 1):
        sheet.cell(row=1, column=column).fill = TITLE_FILL
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(width, 1))
    sheet.row_dimensions[1].height = 24
    return 3


def _header_row(sheet, row: int, headers: Sequence[str]) -> int:
    for column, text in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BOX
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)
    return row + 1


def _widths(sheet, widths: Sequence[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_row(sheet, row: int, values: Sequence[Any], *, wrap: bool = True,
               fills: Optional[Dict[int, PatternFill]] = None) -> int:
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=column, value=value)
        cell.alignment = WRAP if wrap else TOP
        cell.border = BOX
        if fills and column in fills:
            cell.fill = fills[column]
    return row + 1


def _note(sheet, row: int, text: str, width: int = 1) -> int:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.alignment = WRAP
    cell.font = Font(italic=True, color="595959")
    if width > 1:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    return row + 1


# --------------------------------------------------------------------------- #
# Documentation-based sheets
# --------------------------------------------------------------------------- #

def sheet_overview(workbook: Workbook, measured: Optional[Dict[str, Any]]) -> None:
    sheet = workbook.active
    sheet.title = "Overview"
    _widths(sheet, [34, 96])
    row = _title(sheet, "Payment Gateway API Benchmark", 2)

    row = _write_row(sheet, row, ["Built", datetime.now(timezone.utc)
                                  .strftime("%Y-%m-%d %H:%M UTC")])
    if measured:
        meta = measured.get("meta", {})
        row = _write_row(sheet, row, ["Timing data", "PRESENT - measured values below"])
        row = _write_row(sheet, row, ["Measured at", meta.get("generated_at_utc", "?")])
        row = _write_row(sheet, row, ["Measured from",
                                      meta.get("measurement_location_note", "?")])
        row = _write_row(sheet, row, ["Runs per flow",
                                      f"{meta.get('runs_per_flow', '?')} "
                                      f"({meta.get('warmup_runs', '?')} warmups discarded)"])
        row = _write_row(sheet, row, ["Amount / currency",
                                      f"{meta.get('amount', '?')} {meta.get('currency', '?')}"])
        row = _write_row(sheet, row, ["Gateways measured",
                                      ", ".join(meta.get("gateways_run", [])) or "none"])
        row = _write_row(sheet, row, ["Not configured",
                                      ", ".join(meta.get("gateways_not_configured", [])) or "none"])
    else:
        row = _write_row(sheet, row, [
            "Timing data",
            "NOT PRESENT - every timing cell below is empty and awaiting a live run. "
            "Export measured results from BuraPay as JSON, then re-run "
            "scripts/build_workbook.py to fold the numbers in."])

    row += 1
    cell = sheet.cell(row=row, column=1, value="Methodology")
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    sheet.cell(row=row, column=2).fill = SUBHEADER_FILL
    row += 1
    for item in METHODOLOGY:
        row = _write_row(sheet, row, ["", item])

    row += 1
    cell = sheet.cell(row=row, column=1, value="Key findings (documentation-based)")
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    sheet.cell(row=row, column=2).fill = SUBHEADER_FILL
    row += 1
    for item in KEY_FINDINGS:
        row = _write_row(sheet, row, ["", item])

    row += 1
    cell = sheet.cell(row=row, column=1, value="Tabs")
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    sheet.cell(row=row, column=2).fill = SUBHEADER_FILL
    row += 1
    for name, description in [
        ("Call Counts (Documented)", "Documented minimum merchant-server calls per flow."),
        ("Endpoint Map", "Every endpoint each gateway uses, per flow."),
        ("Timing Results", "End-to-end latency per gateway+flow. Empty until a live run."),
        ("Per-Call Timings", "Latency of each individual call in each flow."),
        ("Documented vs Measured", "Where the sandbox disagrees with the docs."),
        ("Data Gaps", "The 22 claims that need live sandbox confirmation."),
        ("Sources", "Documentation URLs behind every claim."),
    ]:
        row = _write_row(sheet, row, [name, description])


def sheet_call_counts(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Call Counts (Documented)")
    _widths(sheet, [22] + [30] * len(GATEWAYS))
    row = _title(sheet, "Documented minimum merchant-server API calls",
                 len(GATEWAYS) + 1)
    row = _header_row(sheet, row, ["Flow"] + [g["label"] for g in GATEWAYS])

    baseline_column = 2  # Geidea
    for flow in FLOWS:
        values = [FLOW_LABELS[flow]] + [CALL_COUNT_MATRIX[g["name"]][flow]
                                        for g in GATEWAYS]
        fills = {baseline_column: BASELINE_FILL}
        for index, gateway in enumerate(GATEWAYS, start=2):
            if "unverified" in CALL_COUNT_MATRIX[gateway["name"]][flow].lower():
                fills[index] = FLAG_FILL
        row = _write_row(sheet, row, values, fills=fills)

    row += 1
    row = _note(sheet, row,
                "Geidea (highlighted) is the baseline. Orange cells rest on an inferred or "
                "unconfirmed source - see the Data Gaps tab.", len(GATEWAYS) + 1)
    _note(sheet, row,
          "Counts are outbound merchant-server calls only. Inbound webhooks are noted in "
          "the cell text but not counted, since the merchant does not initiate them and "
          "cannot time them as a round trip.", len(GATEWAYS) + 1)


def sheet_endpoint_map(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Endpoint Map")
    _widths(sheet, [16, 20, 10, 58, 46])
    row = _title(sheet, "Endpoint map - the exact calls each flow makes", 5)
    row = _header_row(sheet, row, ["Gateway", "Flow", "Step", "Endpoint", "Note"])

    for gateway in GATEWAYS:
        for flow in FLOWS:
            for step in ENDPOINT_MAP[gateway["name"]].get(flow, []):
                fills = {5: FLAG_FILL} if step.get("flagged") else None
                row = _write_row(sheet, row, [
                    gateway["label"], FLOW_LABELS[flow], step["step"],
                    step["endpoint"], step.get("note", "")], fills=fills)


def sheet_data_gaps(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Data Gaps")
    _widths(sheet, [5, 16, 70, 46, 14])
    row = _title(sheet, "Open questions requiring live sandbox confirmation", 5)
    row = _header_row(sheet, row, ["#", "Gateway", "Gap", "What resolves it", "Status"])
    for gap in DATA_GAPS:
        row = _write_row(sheet, row,
                         [gap["id"], gap["gateway"], gap["gap"], gap["resolution"], "OPEN"],
                         fills={5: FLAG_FILL})
    row += 1
    _note(sheet, row,
          "Update Status as each is confirmed or refuted against a real sandbox. "
          "Gaps 2 (Geidea frictionless shortcut) and 16 (HyperPay Direct API call count) "
          "are the two that most affect the headline comparison.", 5)


def sheet_sources(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Sources")
    _widths(sheet, [16, 40, 76])
    row = _title(sheet, "Documentation sources", 3)
    row = _header_row(sheet, row, ["Gateway", "What it covers", "URL"])
    for source in SOURCES:
        row = _write_row(sheet, row, [source["gateway"], source["covers"], source["url"]],
                         wrap=False)
    row += 1
    _note(sheet, row, "All pages fetched 2026-08-18 unless noted otherwise.", 3)


# --------------------------------------------------------------------------- #
# Measurement sheets
# --------------------------------------------------------------------------- #

TIMING_HEADERS = ["Gateway", "Flow", "Status", "Runs OK", "API calls", "min ms", "mean ms",
                  "median ms", "p95 ms", "max ms", "stdev ms", "Note"]


def sheet_timing(workbook: Workbook, measured: Optional[Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Timing Results")
    _widths(sheet, [16, 20, 12, 10, 11, 11, 11, 11, 11, 11, 11, 52])
    row = _title(sheet, "End-to-end latency per flow (milliseconds)", len(TIMING_HEADERS))
    row = _header_row(sheet, row, TIMING_HEADERS)

    results = (measured or {}).get("results", [])
    if not results:
        # Leave the grid pre-labelled so a reader can see exactly what is missing,
        # rather than facing an empty sheet.
        for gateway in GATEWAYS:
            for flow in FLOWS:
                row = _write_row(sheet, row,
                                 [gateway["label"], FLOW_LABELS[flow], "awaiting run",
                                  "", "", "", "", "", "", "", "", ""])
        _note(sheet, row + 1,
              "Empty by design. Run benchmarks in BuraPay against live sandboxes, then "
              "scripts/build_workbook.py, to populate these rows.", len(TIMING_HEADERS))
        return

    labels = {g["name"]: g["label"] for g in GATEWAYS}
    for result in results:
        e2e = result["end_to_end"]
        status = "skipped" if result["skipped"] else ("ok" if result["runs_ok"] else "FAILED")
        note = result.get("skip_reason") or (result["errors"][-1][:250]
                                             if result.get("errors") else "")
        row = _write_row(sheet, row, [
            labels.get(result["gateway"], result["gateway"]),
            FLOW_LABELS.get(result["flow"], result["flow"]),
            status, f"{result['runs_ok']}/{result['runs_attempted']}",
            result["call_count"], e2e["min"], e2e["mean"], e2e["median"],
            e2e["p95"], e2e["max"], e2e["stdev"], note],
            fills={3: FLAG_FILL} if status != "ok" else None)


def sheet_per_call(workbook: Workbook, measured: Optional[Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Per-Call Timings")
    headers = ["Gateway", "Flow", "Step", "Call", "Method", "Samples", "min ms", "mean ms",
               "median ms", "p95 ms", "max ms", "stdev ms"]
    _widths(sheet, [16, 20, 7, 40, 9, 10, 11, 11, 11, 11, 11, 11])
    row = _title(sheet, "Latency of each individual call (milliseconds)", len(headers))
    row = _header_row(sheet, row, headers)

    results = (measured or {}).get("results", [])
    if not results:
        _note(sheet, row,
              "Empty by design - populated from results/latest.json after a live run.",
              len(headers))
        return

    labels = {g["name"]: g["label"] for g in GATEWAYS}
    for result in results:
        for call in result.get("per_call", []):
            row = _write_row(sheet, row, [
                labels.get(result["gateway"], result["gateway"]),
                FLOW_LABELS.get(result["flow"], result["flow"]),
                call["step"], call["name"], call["method"], call["n"],
                call["min"], call["mean"], call["median"], call["p95"],
                call["max"], call["stdev"]], wrap=False)


def sheet_doc_vs_measured(workbook: Workbook, measured: Optional[Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Documented vs Measured")
    headers = ["Gateway", "Flow", "Documented minimum", "Measured calls", "Agrees?"]
    _widths(sheet, [16, 20, 46, 15, 12])
    row = _title(sheet, "Does the sandbox behave the way the docs say?", len(headers))
    row = _header_row(sheet, row, headers)

    results = (measured or {}).get("results", [])
    documented = (measured or {}).get("meta", {}).get("documented_calls", {})
    labels = {g["name"]: g["label"] for g in GATEWAYS}

    if not results:
        for gateway in GATEWAYS:
            for flow in FLOWS:
                row = _write_row(sheet, row, [
                    gateway["label"], FLOW_LABELS[flow],
                    CALL_COUNT_MATRIX[gateway["name"]][flow], "", "awaiting run"])
        _note(sheet, row + 1,
              "The middle column is what each vendor documents. The right-hand columns "
              "fill in once a live run has happened - a mismatch in either direction is "
              "a finding worth writing up.", len(headers))
        return

    for result in results:
        if result["skipped"] or not result["runs_ok"]:
            continue
        gateway = result["gateway"]
        doc_text = documented.get(gateway, {}).get(result["flow"],
                                                   CALL_COUNT_MATRIX.get(gateway, {})
                                                   .get(result["flow"], "-"))
        measured_count = result["call_count"]
        agrees = "yes" if str(measured_count) in doc_text else "CHECK"
        row = _write_row(sheet, row, [
            labels.get(gateway, gateway), FLOW_LABELS.get(result["flow"], result["flow"]),
            doc_text, measured_count, agrees],
            fills={5: FLAG_FILL} if agrees == "CHECK" else None)

    row += 1
    _note(sheet, row,
          "'CHECK' means the measured count is not the documented minimum. That is not "
          "automatically an error - a conditional 3DS step or a skipped optional call "
          "moves the number legitimately - but each one deserves an explanation.",
          len(headers))


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #

def load_measured(path: Optional[str]) -> Optional[Dict[str, Any]]:
    if path is None:
        return None
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as handle:
        return json.load(handle)


def build(output: str, results_path: Optional[str]) -> str:
    measured = load_measured(results_path)
    workbook = Workbook()
    sheet_overview(workbook, measured)
    sheet_call_counts(workbook)
    sheet_endpoint_map(workbook)
    sheet_timing(workbook, measured)
    sheet_per_call(workbook, measured)
    sheet_doc_vs_measured(workbook, measured)
    sheet_data_gaps(workbook)
    sheet_sources(workbook)
    os.makedirs(os.path.dirname(output), exist_ok=True)
    workbook.save(output)
    return output


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--results", default=os.path.join(RESULTS_DIR, "latest.json"),
                        help="Measured results JSON (default: ../results/latest.json)")
    parser.add_argument("--docs-only", action="store_true",
                        help="Build the documentation-based workbook, ignoring measured data")
    parser.add_argument("--output",
                        default=os.path.join(RESULTS_DIR, "gateway_benchmark_workbook.xlsx"),
                        help="Where to write the workbook")
    args = parser.parse_args(argv)

    results_path = None if args.docs_only else args.results
    if results_path and not os.path.exists(results_path):
        print(f"No measured results at {results_path} - building the documentation-based "
              f"workbook with empty timing tabs.")
        results_path = None

    path = build(args.output, results_path)
    print(f"Wrote {path}")
    if results_path is None:
        print("Timing tabs are empty. Run benchmarks in BuraPay, export the results as "
              "JSON, then re-run "
              "this script to fold the numbers in.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
