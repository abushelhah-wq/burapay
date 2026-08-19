"""
Exports (specification section 28): CSV, Excel and JSON.

Exports carry the same caveats the UI does. Every file starts with the filter set and
the sample size that produced it, because a CSV of latencies with no sample size is a
number without a claim attached — and these files are exactly what ends up pasted
into a decision document.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(color="FFFFFF", bold=True)

TRANSACTION_COLUMNS: Sequence[str] = (
    "id", "started_at", "completed_at", "gateway_code", "integration_type",
    "environment", "status", "amount", "currency", "merchant_reference",
    "gateway_transaction_id", "api_call_count", "gateway_api_time_ms",
    "three_ds_time_ms", "customer_interaction_time_ms", "redirect_time_ms",
    "page_load_time_ms", "total_duration_ms", "app_overhead_ms", "webhook_latency_ms",
    "error_category", "gateway_response_code", "gateway_response_message",
    "methodology", "benchmark_run_id",
)


def _value(row: Any, column: str) -> Any:
    value = getattr(row, column, None)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def transactions_to_rows(transactions: Iterable[Any]) -> List[Dict[str, Any]]:
    return [{column: _value(t, column) for column in TRANSACTION_COLUMNS}
            for t in transactions]


def to_csv(rows: Sequence[Dict[str, Any]], *, columns: Sequence[str] | None = None) -> bytes:
    buffer = io.StringIO()
    fieldnames = list(columns or (rows[0].keys() if rows else TRANSACTION_COLUMNS))
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")      # BOM so Excel reads UTF-8


def to_json(payload: Dict[str, Any]) -> bytes:
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def _write_sheet(workbook: Workbook, title: str, rows: Sequence[Dict[str, Any]],
                 columns: Sequence[str] | None = None) -> None:
    sheet = workbook.create_sheet(title[:31])
    fieldnames = list(columns or (rows[0].keys() if rows else ["(no data)"]))
    sheet.append(fieldnames)
    for index, name in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
        sheet.column_dimensions[get_column_letter(index)].width = max(14, min(38, len(name) + 6))
    for row in rows:
        sheet.append([row.get(name) for name in fieldnames])
    sheet.freeze_panes = "A2"


def to_excel(sheets: Dict[str, Sequence[Dict[str, Any]]], *,
             context: Dict[str, Any] | None = None) -> bytes:
    """A workbook with one sheet per dataset, plus an About sheet naming the filters."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    about = workbook.create_sheet("About")
    about.append(["BuraPay Gateway Benchmark — export"])
    about["A1"].font = Font(bold=True, size=14)
    about.append(["Generated", datetime.now(timezone.utc).isoformat()])
    for key, value in (context or {}).items():
        about.append([key, json.dumps(value, default=str) if isinstance(value, (dict, list))
                      else value])
    about.append([])
    about.append(["Note", "Latency figures cover completed transactions only. Sample "
                          "sizes are included with every group; a percentile over a "
                          "small sample is not a reliable figure."])
    about.column_dimensions["A"].width = 34
    about.column_dimensions["B"].width = 90

    for title, rows in sheets.items():
        _write_sheet(workbook, title, rows)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def flatten_summary(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Flatten the nested comparison rows into one spreadsheet-friendly row each."""
    flat: List[Dict[str, Any]] = []
    for row in rows:
        api, total = row.get("api", {}) or {}, row.get("total", {}) or {}
        flat.append({
            "gateway_code": row.get("gateway_code"),
            "gateway_name": row.get("gateway_name"),
            "integration_type": row.get("integration_type"),
            "transactions": row.get("transactions"),
            "completed": row.get("completed"),
            "success_rate_pct": row.get("success_rate"),
            "failure_rate_pct": row.get("failure_rate"),
            "api_count": api.get("count"),
            "api_min_ms": api.get("min"),
            "api_mean_ms": api.get("mean"),
            "api_median_ms": api.get("median"),
            "api_p90_ms": api.get("p90"),
            "api_p95_ms": api.get("p95"),
            "api_p99_ms": api.get("p99"),
            "api_max_ms": api.get("max"),
            "api_stdev_ms": api.get("stdev"),
            "api_sample_reliable": api.get("reliable"),
            "total_mean_ms": total.get("mean"),
            "total_p95_ms": total.get("p95"),
            "api_call_count": row.get("api_call_count"),
            "documented_calls": row.get("documented_calls"),
            "is_simulated": row.get("is_simulated"),
        })
    return flat
